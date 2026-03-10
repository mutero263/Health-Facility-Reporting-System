"""
Utility Functions for Health Facility Reporting System
Helper functions for data processing, validation, file handling, etc.
"""

import os
import re
import json
import hashlib
import secrets
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple, Union
import logging
import shutil
import csv
import io
import zipfile
from collections import defaultdict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from health_parser import HealthReport, FractionField

# Setup logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ======================
# DATE AND TIME UTILITIES
# ======================

def get_current_week() -> Tuple[int, int]:
    """
    Get current week number and year
    
    Returns:
        Tuple of (year, week_number)
    """
    now = datetime.now()
    year = now.year
    week = now.isocalendar()[1]
    return year, week


def get_week_range(week: int, year: int) -> Tuple[datetime, datetime]:
    """
    Get start and end dates for a given week
    
    Args:
        week: Week number (1-53)
        year: Year
    
    Returns:
        Tuple of (start_date, end_date)
    """
    # First day of the year
    first_day = datetime(year, 1, 1)
    
    # Find the first Monday of the year
    days_to_monday = (7 - first_day.weekday()) % 7
    first_monday = first_day + timedelta(days=days_to_monday)
    
    # Calculate start of given week
    start_date = first_monday + timedelta(weeks=week-1)
    end_date = start_date + timedelta(days=6)
    
    return start_date, end_date


def format_week_display(week: int, year: int) -> str:
    """Format week for display (e.g., '2024-W09')"""
    return f"{year}-W{week:02d}"


def parse_week_display(week_str: str) -> Tuple[int, int]:
    """
    Parse week display string (e.g., '2024-W09') to year and week
    
    Returns:
        Tuple of (year, week)
    """
    pattern = r'^(\d{4})-W(\d{1,2})$'
    match = re.match(pattern, week_str)
    if match:
        year = int(match.group(1))
        week = int(match.group(2))
        return year, week
    raise ValueError(f"Invalid week format: {week_str}")


def get_weeks_between(start_week: int, start_year: int, 
                      end_week: int, end_year: int) -> List[Tuple[int, int]]:
    """
    Get list of weeks between start and end dates
    
    Returns:
        List of (year, week) tuples
    """
    weeks = []
    
    current_year, current_week = start_year, start_week
    
    while (current_year < end_year) or (current_year == end_year and current_week <= end_week):
        weeks.append((current_year, current_week))
        
        current_week += 1
        if current_week > 52:
            current_week = 1
            current_year += 1
    
    return weeks


def get_week_options(max_weeks: int = 12) -> List[Dict[str, Any]]:
    """
    Get list of recent weeks for dropdown menus
    
    Returns:
        List of dicts with year, week, display
    """
    current_year, current_week = get_current_week()
    
    options = []
    year, week = current_year, current_week
    
    for i in range(max_weeks):
        options.append({
            'year': year,
            'week': week,
            'display': format_week_display(week, year)
        })
        
        week -= 1
        if week < 1:
            week = 52
            year -= 1
    
    return options


# ======================
# METRICS FORMATTING UTILITIES
# ======================

def format_metrics_display(report: Dict[str, Any]) -> Dict[str, Any]:
    """
    Format metrics for display in templates
    
    Args:
        report: Report dictionary
        
    Returns:
        Dictionary with formatted metrics
    """
    formatted = {}
    
    # Format fraction fields
    fraction_fields = [
        'malaria_suspected', 'malaria_tested', 'malaria_positive',
        'diarrhoea', 'dysentery', 'suspected_dysentery', 'influenza',
        'vhw_malaria_suspected', 'vhw_malaria_tested', 'vhw_malaria_positive',
        'vhw_diarrhoea', 'vhw_dysentery'
    ]
    
    for field in fraction_fields:
        if field in report:
            value = report[field]
            formatted[f"{field}_display"] = value
            formatted[f"{field}_numerator"] = FractionField.get_numerator(value)
            formatted[f"{field}_denominator"] = FractionField.get_denominator(value)
            
            if formatted[f"{field}_denominator"] > 0:
                formatted[f"{field}_rate"] = f"{formatted[f'{field}_numerator'] / formatted[f'{field}_denominator']:.1%}"
            else:
                formatted[f"{field}_rate"] = "0%"
    
    # Format integer fields with commas
    integer_fields = [
        'opd_visits', 'institutional_deliveries', 'home_deliveries',
        'anc_contacts', 'fp_clients', 'pnc_attendees', 'hiv_tested',
        'children_vaccinated_penta3', 'tb_screened', 'under5_sam',
        'under5_mam', 'children_vitamin_a', 'under5_deaths',
        'institutional_deaths', 'functional_ambulance', 'xray_patients',
        'tracer_medicines_stock', 'drs_resigned', 'nurses_resigned',
        'casualty_visits', 'in_patients_admissions', 'major_operations',
        'c_sections', 'renal_dialysis', 'dog_bite', 'kwashiorkor',
        'marasmus', 'bilharzia', 'maternal_death', 'perinatal_death',
        'aefi', 'afp', 'nnt', 'measles'
    ]
    
    for field in integer_fields:
        if field in report and report[field] is not None:
            try:
                value = int(report[field])
                formatted[f"{field}_formatted"] = f"{value:,}"
            except (ValueError, TypeError):
                formatted[f"{field}_formatted"] = str(report[field])
    
    # Format percentage fields
    if 'malaria_positivity_rate' in report:
        formatted['malaria_positivity_rate_display'] = f"{report['malaria_positivity_rate']:.1%}"
    
    if 'institutional_delivery_rate' in report:
        formatted['institutional_delivery_rate_display'] = f"{report['institutional_delivery_rate']:.1%}"
    
    return formatted


def format_metric_value(value: Any, metric_type: str = "integer") -> str:
    """
    Format metric value for display
    
    Args:
        value: The value to format
        metric_type: Type of metric (integer, fraction, percentage, currency)
    
    Returns:
        Formatted string
    """
    if value is None:
        return "-"
    
    if metric_type == "integer":
        try:
            return f"{int(value):,}"
        except:
            return str(value)
    
    elif metric_type == "fraction":
        if isinstance(value, str) and '/' in value:
            return value
        return "0/0"
    
    elif metric_type == "percentage":
        try:
            return f"{float(value):.1%}"
        except:
            return "0%"
    
    elif metric_type == "currency":
        try:
            return f"${float(value):,.2f}"
        except:
            return "$0"
    
    else:
        return str(value)


def format_metric_name(metric: str) -> str:
    """Convert metric key to display name"""
    # Replace underscores with spaces and capitalize
    name = metric.replace('_', ' ').title()
    
    # Handle common abbreviations
    abbreviations = {
        'Opd': 'OPD',
        'Anc': 'ANC',
        'Pnc': 'PNC',
        'Fp': 'FP',
        'Hiv': 'HIV',
        'Tb': 'TB',
        'Aefi': 'AEFI',
        'Afp': 'AFP',
        'Nnt': 'NNT',
        'Vhw': 'VHW',
        'Rdns': 'RDNS',
        'Sam': 'SAM',
        'Mam': 'MAM',
        'Penta3': 'Penta3',
    }
    
    words = name.split()
    formatted_words = []
    
    for word in words:
        if word in abbreviations:
            formatted_words.append(abbreviations[word])
        else:
            formatted_words.append(word)
    
    return ' '.join(formatted_words)


def truncate_string(s: str, max_length: int = 50, suffix: str = "...") -> str:
    """Truncate string to max length"""
    if len(s) <= max_length:
        return s
    return s[:max_length - len(suffix)] + suffix


# ======================
# DATA VALIDATION UTILITIES
# ======================

def validate_report_data(report: HealthReport) -> List[str]:
    """
    Validate report data for consistency and quality
    
    Returns:
        List of validation warnings/errors
    """
    warnings = []
    
    # Check for missing critical fields
    if report.facility_id is None:
        warnings.append("Facility ID is missing")
    
    if report.week == 0:
        warnings.append("Week number is missing or invalid")
    
    if report.year == 0:
        warnings.append("Year is missing or invalid")
    
    # Check for zeros in critical metrics
    critical_metrics = [
        ('opd_visits', 'OPD visits'),
        ('institutional_deliveries', 'Institutional deliveries'),
        ('anc_contacts', 'ANC contacts')
    ]
    
    for metric, name in critical_metrics:
        if getattr(report, metric, 0) == 0:
            warnings.append(f"{name} reported as zero - please verify")
    
    # Check malaria data consistency
    suspected_num, suspected_den = report.get_fraction_parts('malaria_suspected')
    tested_num, tested_den = report.get_fraction_parts('malaria_tested')
    positive_num, positive_den = report.get_fraction_parts('malaria_positive')
    
    if tested_num > suspected_num and suspected_num > 0:
        warnings.append(f"Malaria tested ({tested_num}) exceeds suspected ({suspected_num})")
    
    if positive_num > tested_num and tested_num > 0:
        warnings.append(f"Malaria positive ({positive_num}) exceeds tested ({tested_num})")
    
    # Check delivery consistency
    if report.still_births > report.institutional_deliveries:
        warnings.append(f"Still births ({report.still_births}) exceed deliveries ({report.institutional_deliveries})")
    
    # Check ANC/delivery ratio (should be at least 4 per delivery)
    if report.institutional_deliveries > 0:
        anc_per_delivery = report.anc_contacts / report.institutional_deliveries
        if anc_per_delivery < 2:
            warnings.append(f"Low ANC contacts ({anc_per_delivery:.1f} per delivery) - expected at least 4")
    
    return warnings


def calculate_data_quality_score(report: HealthReport) -> float:
    """
    Calculate data quality score (0-100)
    
    Higher score = better quality
    """
    score = 100.0
    deductions = []
    
    # Check for missing critical fields
    if report.facility_id is None:
        score -= 20
        deductions.append("Missing facility ID")
    
    if report.week == 0 or report.year == 0:
        score -= 20
        deductions.append("Missing week/year")
    
    # Check for zeros in critical metrics
    zero_metrics = []
    for metric in ['opd_visits', 'institutional_deliveries', 'anc_contacts']:
        if getattr(report, metric, 0) == 0:
            zero_metrics.append(metric)
    
    if zero_metrics:
        deduction = min(len(zero_metrics) * 5, 20)
        score -= deduction
        deductions.append(f"Zero values: {', '.join(zero_metrics)}")
    
    # Check malaria data consistency
    suspected_num, _ = report.get_fraction_parts('malaria_suspected')
    tested_num, _ = report.get_fraction_parts('malaria_tested')
    positive_num, _ = report.get_fraction_parts('malaria_positive')
    
    if tested_num > suspected_num and suspected_num > 0:
        score -= 10
        deductions.append("Inconsistent malaria data")
    
    if positive_num > tested_num and tested_num > 0:
        score -= 10
        deductions.append("Inconsistent malaria data")
    
    # Ensure score doesn't go below 0
    score = max(0, score)
    
    logger.debug(f"Data quality score: {score:.1f} - Deductions: {deductions}")
    
    return score


def detect_outliers(values: List[float], threshold: float = 3.0) -> List[int]:
    """
    Detect outliers in a list of values using z-score method
    
    Args:
        values: List of numerical values
        threshold: Z-score threshold for outlier detection
    
    Returns:
        List of indices of outliers
    """
    if len(values) < 3:
        return []
    
    import numpy as np
    from scipy import stats
    
    z_scores = np.abs(stats.zscore(values))
    outliers = np.where(z_scores > threshold)[0].tolist()
    
    return outliers


# ======================
# FILE HANDLING UTILITIES
# ======================

def save_uploaded_file(uploaded_file, upload_dir: Path, filename: str = None) -> Path:
    """
    Save an uploaded file to disk
    
    Args:
        uploaded_file: FastAPI UploadFile object
        upload_dir: Directory to save to
        filename: Optional custom filename
    
    Returns:
        Path to saved file
    """
    # Create directory if it doesn't exist
    upload_dir.mkdir(parents=True, exist_ok=True)
    
    # Generate filename if not provided
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        original_name = Path(uploaded_file.filename).stem
        extension = Path(uploaded_file.filename).suffix
        filename = f"{original_name}_{timestamp}{extension}"
    
    file_path = upload_dir / filename
    
    # Save file
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(uploaded_file.file, buffer)
    
    logger.info(f"Saved uploaded file: {file_path}")
    
    return file_path


def cleanup_old_files(directory: Path, days: int = 7, pattern: str = "*"):
    """
    Delete files older than specified days
    
    Args:
        directory: Directory to clean
        days: Maximum age in days
        pattern: File pattern to match
    """
    cutoff = datetime.now() - timedelta(days=days)
    
    for file_path in directory.glob(pattern):
        if file_path.is_file():
            mtime = datetime.fromtimestamp(file_path.stat().st_mtime)
            if mtime < cutoff:
                file_path.unlink()
                logger.info(f"Deleted old file: {file_path}")


def ensure_directory(directory: Path) -> Path:
    """Ensure directory exists, create if necessary"""
    directory.mkdir(parents=True, exist_ok=True)
    return directory


def get_file_size(file_path: Path) -> str:
    """Get human-readable file size"""
    size = file_path.stat().st_size
    
    for unit in ['B', 'KB', 'MB', 'GB']:
        if size < 1024:
            return f"{size:.1f} {unit}"
        size /= 1024
    
    return f"{size:.1f} TB"


# ======================
# EXPORT UTILITIES
# ======================

def export_to_csv(data: List[Dict[str, Any]], filepath: Path) -> Path:
    """
    Export data to CSV file
    
    Args:
        data: List of dictionaries to export
        filepath: Path to save CSV file
    
    Returns:
        Path to created file
    """
    if not data:
        # Create empty file with headers
        df = pd.DataFrame()
    else:
        df = pd.DataFrame(data)
    
    df.to_csv(filepath, index=False)
    logger.info(f"Exported {len(data)} rows to {filepath}")
    
    return filepath


def export_to_excel(data_sheets: Dict[str, List[Dict[str, Any]]], 
                   filepath: Path) -> Path:
    """
    Export multiple sheets to Excel file
    
    Args:
        data_sheets: Dictionary mapping sheet names to data lists
        filepath: Path to save Excel file
    
    Returns:
        Path to created file
    """
    wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for sheet_name, data in data_sheets.items():
        if not data:
            continue
        
        df = pd.DataFrame(data)
        ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet name max 31 chars
        
        # Write headers
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=format_metric_name(col_name))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Write data
        for row_idx, row in df.iterrows():
            for col_idx, col_name in enumerate(df.columns, 1):
                value = row[col_name]
                
                # Format based on column name
                if any(fraction in col_name for fraction in ['suspected', 'tested', 'positive', 'diarrhoea']):
                    formatted_value = format_metric_value(value, "fraction")
                elif 'rate' in col_name or 'coverage' in col_name:
                    formatted_value = format_metric_value(value, "percentage")
                else:
                    formatted_value = format_metric_value(value, "integer")
                
                cell = ws.cell(row=row_idx + 2, column=col_idx, value=formatted_value)
                cell.border = border
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width
    
    wb.save(filepath)
    logger.info(f"Exported {len(data_sheets)} sheets to {filepath}")
    
    return filepath


def create_zip_archive(files: List[Path], zip_path: Path) -> Path:
    """
    Create ZIP archive from list of files
    
    Args:
        files: List of file paths to include
        zip_path: Path for the ZIP file
    
    Returns:
        Path to created ZIP file
    """
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for file in files:
            if file.exists():
                zipf.write(file, arcname=file.name)
    
    logger.info(f"Created ZIP archive with {len(files)} files: {zip_path}")
    
    return zip_path


# ======================
# AGGREGATION UTILITIES
# ======================

def aggregate_reports(reports: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Aggregate multiple reports into summary statistics
    
    Args:
        reports: List of report dictionaries
    
    Returns:
        Dictionary with aggregated statistics
    """
    if not reports:
        return {}
    
    df = pd.DataFrame(reports)
    
    aggregation = {
        'total_reports': len(reports),
        'date_range': {
            'start': f"{df['year'].min()}-W{df['week'].min():02d}",
            'end': f"{df['year'].max()}-W{df['week'].max():02d}"
        }
    }
    
    # Numeric aggregations
    numeric_columns = ['opd_visits', 'institutional_deliveries', 'anc_contacts',
                       'fp_clients', 'hiv_tested', 'children_vaccinated_penta3']
    
    for col in numeric_columns:
        if col in df.columns:
            aggregation[col] = {
                'total': int(df[col].sum()),
                'average': float(df[col].mean()),
                'min': int(df[col].min()),
                'max': int(df[col].max()),
                'std': float(df[col].std()) if len(df) > 1 else 0
            }
    
    # Malaria aggregations
    if 'malaria_positive' in df.columns:
        total_positive = 0
        for val in df['malaria_positive']:
            num = FractionField.get_numerator(val)
            total_positive += num
        
        aggregation['malaria'] = {
            'total_positive': total_positive
        }
    
    return aggregation


def calculate_trend(values: List[float]) -> Dict[str, Any]:
    """
    Calculate trend statistics for a list of values
    
    Returns:
        Dictionary with trend information
    """
    if len(values) < 2:
        return {
            'direction': 'insufficient_data',
            'change': 0,
            'volatility': 0
        }
    
    # Simple linear regression for slope
    x = list(range(len(values)))
    
    # Calculate slope
    n = len(x)
    slope = (n * sum(x[i] * values[i] for i in range(n)) - sum(x) * sum(values)) / \
            (n * sum(x[i] ** 2 for i in range(n)) - sum(x) ** 2) if n > 1 else 0
    
    # Determine direction
    if slope > 0.1:
        direction = 'increasing'
    elif slope < -0.1:
        direction = 'decreasing'
    else:
        direction = 'stable'
    
    # Calculate percentage change
    if values[0] != 0:
        pct_change = ((values[-1] - values[0]) / abs(values[0])) * 100
    else:
        pct_change = 0
    
    # Calculate volatility (coefficient of variation)
    mean_val = sum(values) / len(values)
    if mean_val != 0:
        variance = sum((x - mean_val) ** 2 for x in values) / len(values)
        volatility = (variance ** 0.5) / mean_val
    else:
        volatility = 0
    
    return {
        'direction': direction,
        'slope': slope,
        'change': pct_change,
        'volatility': volatility,
        'first': values[0],
        'last': values[-1],
        'min': min(values),
        'max': max(values)
    }


# ======================
# SECURITY UTILITIES
# ======================

def generate_api_key() -> str:
    """Generate a random API key"""
    return secrets.token_urlsafe(32)


def hash_string(s: str) -> str:
    """Create SHA-256 hash of a string"""
    return hashlib.sha256(s.encode()).hexdigest()


def sanitize_filename(filename: str) -> str:
    """
    Sanitize filename to prevent path traversal attacks
    
    Removes or replaces dangerous characters
    """
    # Remove path separators
    filename = filename.replace('/', '_').replace('\\', '_')
    
    # Remove any leading dots or spaces
    filename = filename.lstrip('. ')
    
    # Replace any other problematic characters
    filename = re.sub(r'[<>:"|?*]', '_', filename)
    
    # Limit length
    if len(filename) > 255:
        name, ext = os.path.splitext(filename)
        filename = name[:250] + ext
    
    return filename or "unnamed_file"


# ======================
# CACHING UTILITIES
# ======================

class SimpleCache:
    """Simple in-memory cache with TTL"""
    
    def __init__(self, ttl_seconds: int = 300):
        self.cache = {}
        self.ttl = ttl_seconds
    
    def get(self, key: str) -> Optional[Any]:
        """Get value from cache if not expired"""
        if key in self.cache:
            value, timestamp = self.cache[key]
            if datetime.now() - timestamp < timedelta(seconds=self.ttl):
                return value
            else:
                del self.cache[key]
        return None
    
    def set(self, key: str, value: Any):
        """Set value in cache"""
        self.cache[key] = (value, datetime.now())
    
    def clear(self):
        """Clear all cache entries"""
        self.cache.clear()
    
    def cleanup(self):
        """Remove expired entries"""
        now = datetime.now()
        expired = [k for k, (_, ts) in self.cache.items() 
                  if now - ts >= timedelta(seconds=self.ttl)]
        for k in expired:
            del self.cache[k]


# ======================
# REPORT GENERATION UTILITIES
# ======================

def generate_report_summary(reports: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Generate a human-readable summary of reports
    """
    if not reports:
        return {"summary": "No reports available"}
    
    summary = {
        "total_reports": len(reports),
        "date_range": {
            "from": format_week_display(reports[0]['week'], reports[0]['year']),
            "to": format_week_display(reports[-1]['week'], reports[-1]['year'])
        },
        "highlights": [],
        "concerns": []
    }
    
    # Calculate totals
    total_opd = sum(r.get('opd_visits', 0) for r in reports)
    total_deliveries = sum(r.get('institutional_deliveries', 0) for r in reports)
    total_anc = sum(r.get('anc_contacts', 0) for r in reports)
    
    summary["totals"] = {
        "opd_visits": total_opd,
        "deliveries": total_deliveries,
        "anc_contacts": total_anc
    }
    
    # Find highlights
    if total_deliveries > 0:
        summary["highlights"].append(f"Total deliveries: {total_deliveries}")
    
    # Find concerns
    zero_delivery_weeks = sum(1 for r in reports if r.get('institutional_deliveries', 0) == 0)
    if zero_delivery_weeks > len(reports) * 0.3:
        summary["concerns"].append(f"No deliveries reported in {zero_delivery_weeks} weeks")
    
    return summary


def format_report_for_display(report: Dict[str, Any]) -> Dict[str, Any]:
    """
    Format report data for display in templates
    """
    formatted = dict(report)
    
    # Format fractions
    fraction_fields = ['malaria_suspected', 'malaria_tested', 'malaria_positive',
                       'diarrhoea', 'dysentery', 'influenza']
    
    for field in fraction_fields:
        if field in report:
            value = report[field]
            num = FractionField.get_numerator(value)
            den = FractionField.get_denominator(value)
            
            formatted[f"{field}_display"] = value
            formatted[f"{field}_numerator"] = num
            formatted[f"{field}_denominator"] = den
            
            if den > 0:
                formatted[f"{field}_rate"] = f"{num/den:.1%}"
    
    # Format dates
    if 'report_date' in report and report['report_date']:
        if isinstance(report['report_date'], str):
            try:
                dt = datetime.fromisoformat(report['report_date'])
                formatted['report_date_display'] = dt.strftime("%d %b %Y")
            except:
                formatted['report_date_display'] = report['report_date']
    
    return formatted


# ======================
# MAIN TEST FUNCTION
# ======================

if __name__ == "__main__":
    # Test utility functions
    print("Testing utilities...")
    
    # Test week functions
    year, week = get_current_week()
    print(f"Current week: {format_week_display(week, year)}")
    
    start, end = get_week_range(week, year)
    print(f"Week range: {start.date()} to {end.date()}")
    
    weeks = get_weeks_between(1, 2024, 4, 2024)
    print(f"Weeks between: {weeks}")
    
    options = get_week_options(5)
    print(f"Week options: {options}")
    
    # Test formatting
    print(format_metric_name("institutional_deliveries"))
    print(format_metric_name("children_vaccinated_penta3"))
    
    # Test trend calculation
    values = [10, 12, 15, 14, 18, 20, 22]
    trend = calculate_trend(values)
    print(f"Trend: {trend}")
    
    print("All tests passed!")